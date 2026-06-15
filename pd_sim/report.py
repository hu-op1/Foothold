"""Terminal table and XLSX output for simulation results."""

import os
import re

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, numbers
from openpyxl.utils import get_column_letter


def print_comparison_table(results: list[dict], cfg: dict) -> None:
    """Print a comparison table of strategy results.

    Args:
        results: list of {label, metrics_raw, score}
        cfg: simulation config.
    """
    trace_path = cfg["trace"]["path"]
    gpu = cfg.get("gpu", "unknown")
    total_gpus = cfg["strategy"].get("total_gpus", "?")

    print()
    print("=" * 120)
    print("  PD Strategy Comparison")
    if results:
        n_reqs = results[0]["metrics_raw"]["num_requests"]
        print(f"  Trace: {trace_path} ({n_reqs} requests)")
    print(f"  GPU: {gpu} x{total_gpus}")
    print("=" * 120)

    header = (
        f"  {'Strategy':<32} | {'Thrpt':>7} | {'InThrpt':>8} | {'TotalThrpt':>11} | "
        f"{'TTFTmean':>8} | {'TTFTp99':>8} | {'TPOTmean':>9} | {'TPOTp99':>9} | "
        f"{'P99lat':>8} | {'TotTime':>8} | {'SLO%':>6} | {'Time':>6}"
    )
    print(header)
    print("  " + "-" * (len(header) - 2))

    best = max(results, key=lambda r: r["score"]) if results else None
    for entry in results:
        m = entry["metrics_raw"]
        marker = " ★" if entry is best else ""
        tt_s = m.get("total_time_s", 0)
        print(
            f"  {entry['label']:<32} | "
            f"{m['throughput']:>7.0f} | "
            f"{m['input_throughput']:>8.0f} | "
            f"{m['total_throughput']:>11.0f} | "
            f"{m['mean_ttft_ms']:>8.0f} | "
            f"{m['p99_ttft_ms']:>8.0f} | "
            f"{m['mean_tpot_ms']:>9.1f} | "
            f"{m['p99_tpot_ms']:>9.1f} | "
            f"{m['p99_ms']:>8.0f} | "
            f"{tt_s:>8.0f} | "
            f"{entry['slo_score']*100:>5.1f}% | "
            f"{entry['elapsed']:>4.1f}s{marker}"
        )

    if best:
        print()
        print(f"  ★ Optimal: {best['label']}")

    print()


_LABEL_RE = re.compile(r"^(.*?) \(batch=(\d+), thr=(\d+)\)$")


def _parse_label(label: str) -> tuple[str, int, int]:
    """Parse a label into (strategy_type, batch, threshold).

    >>> _parse_label("Colo TP1 DP4 (batch=256, thr=256)")
    ("Colo TP1 DP4", 256, 256)
    >>> _parse_label("Disagg 1P(TP1×1):3D (batch=512, thr=1024)")
    ("Disagg 1P(TP1×1):3D", 512, 1024)
    """
    m = _LABEL_RE.match(label)
    if m:
        return m.group(1), int(m.group(2)), int(m.group(3))
    return label, 0, 0


def export_xlsx(results: list[dict], path: str) -> None:
    """Export results to XLSX with strategy_type, batch, and thr as separate columns."""
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "PD Sim Results"

    # Header styling
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    center_align = Alignment(horizontal="center")

    headers = [
        "strategy_type",
        "batch",
        "thr",
        "throughput_tok_s",
        "input_throughput_tok_s",
        "output_throughput_tok_s",
        "total_throughput_tok_s",
        "ttft_mean_ms", "ttft_p50_ms", "ttft_p90_ms", "ttft_p99_ms",
        "tpot_mean_ms", "tpot_p50_ms", "tpot_p90_ms", "tpot_p99_ms",
        "latency_p50_ms", "latency_p90_ms", "latency_p95_ms", "latency_p99_ms",
        "num_requests",
        "total_input_tokens",
        "total_output_tokens",
        "total_time_s",
        "cache_hit_rate",
        "score",
        "elapsed_s",
    ]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = center_align

    # Data rows
    for row, entry in enumerate(results, 2):
        m = entry["metrics_raw"]
        strategy_type, batch, thr = _parse_label(entry["label"])
        values = [
            strategy_type,
            batch,
            thr,
            m["throughput"],
            m["input_throughput"],
            m["output_throughput"],
            m["total_throughput"],
            m["mean_ttft_ms"], m["p50_ttft_ms"], m["p90_ttft_ms"], m["p99_ttft_ms"],
            m["mean_tpot_ms"], m["p50_tpot_ms"], m["p90_tpot_ms"], m["p99_tpot_ms"],
            m["p50_ms"], m["p90_ms"], m["p95_ms"], m["p99_ms"],
            m["num_requests"],
            m["total_input_tokens"],
            m["total_output_tokens"],
            m["total_time_s"],
            m.get("cache_hit_rate", 0.0),
            entry["score"],
            entry["elapsed"],
        ]
        for col, val in enumerate(values, 1):
            ws.cell(row=row, column=col, value=val)

    # Auto-width columns
    for col in range(1, len(headers) + 1):
        max_width = len(str(headers[col - 1]))
        for row in range(2, len(results) + 2):
            cell_val = str(ws.cell(row=row, column=col).value or "")
            max_width = max(max_width, min(len(cell_val), 40))
        ws.column_dimensions[get_column_letter(col)].width = max_width + 2

    # Freeze header row
    ws.freeze_panes = "A2"

    wb.save(path)
    print(f"Results exported to: {path}")
