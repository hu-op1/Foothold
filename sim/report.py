"""Terminal table and XLSX output for simulation results."""

import os
import re

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, numbers
from openpyxl.utils import get_column_letter


_LABEL_RE = re.compile(r"^(.*?) \(batch=(\d+), thr=(\d+)\)$")


def _parse_label(label: str) -> tuple[str, int, int]:
    """Parse a label into (strategy_type, batch, threshold).

    >>> _parse_label("Colo TP1 DP4 (batch=256, thr=256)")
    ("Colo TP1 DP4", 256, 256)
    >>> _parse_label("Disagg 1P(TP1×1):3D (batch=512, thr=1024)")
    ("Disagg 1P(TP1×1):3D", 512, 1024)
    """
    m = _LABEL_RE.match(label)
    if m:
        return m.group(1), int(m.group(2)), int(m.group(3))
    return label, 0, 0


def _maybe_add_scalability_sheet(wb, results: list[dict]) -> None:
    """Add a 'Scalability' sheet if results contain GPU sweep data."""
    scalability = None
    for r in results:
        if "_scalability" in r:
            scalability = r["_scalability"]
            break

    if not scalability:
        return

    ws = wb.create_sheet("Scalability")

    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    center_align = Alignment(horizontal="center")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    headers = [
        "total_gpus",
        "colo_label",
        "colo_throughput_tok_s",
        "colo_ttft_p99_ms",
        "colo_tpot_p99_ms",
        "colo_p99_latency_ms",
        "colo_slo_score",
        "disagg_label",
        "disagg_throughput_tok_s",
        "disagg_ttft_p99_ms",
        "disagg_tpot_p99_ms",
        "disagg_p99_latency_ms",
        "disagg_slo_score",
        "winner",
    ]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    for row, s in enumerate(scalability, 2):
        colo = s["best_colo"]
        disagg = s["best_disagg"]

        def _m(r, key):
            return r["metrics_raw"].get(key, 0) if r else 0

        ct = _m(colo, "throughput")
        dt = _m(disagg, "throughput")
        winner = "Colocated" if ct >= dt else "Disaggregated"

        values = [
            s["total_gpus"],
            colo["label"] if colo else "—",
            ct,
            _m(colo, "p99_ttft_ms"),
            _m(colo, "p99_tpot_ms"),
            _m(colo, "p99_ms"),
            colo["slo_score"] if colo else 0,
            disagg["label"] if disagg else "—",
            dt,
            _m(disagg, "p99_ttft_ms"),
            _m(disagg, "p99_tpot_ms"),
            _m(disagg, "p99_ms"),
            disagg["slo_score"] if disagg else 0,
            winner,
        ]

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            if isinstance(val, float):
                cell.number_format = '0.000'

        # Highlight winner cell
        winner_cell = ws.cell(row=row, column=len(headers))
        if winner == "Colocated":
            winner_cell.fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE",
                                           fill_type="solid")
        else:
            winner_cell.fill = green_fill

    # Auto-width
    for col in range(1, len(headers) + 1):
        max_width = len(str(headers[col - 1]))
        for row in range(2, len(scalability) + 2):
            cell_val = str(ws.cell(row=row, column=col).value or "")
            max_width = max(max_width, min(len(cell_val), 60))
        ws.column_dimensions[get_column_letter(col)].width = max_width + 2

    ws.freeze_panes = "A2"


def _plot_scalability(results: list[dict], xlsx_path: str) -> None:
    """Generate a scalability plot PNG alongside the xlsx."""
    scalability = None
    for r in results:
        if "_scalability" in r:
            scalability = r["_scalability"]
            break
    if not scalability or len(scalability) < 2:
        return

    gpus = [s["total_gpus"] for s in scalability]
    colo_tp = [s["best_colo"]["metrics_raw"]["throughput"] if s["best_colo"] else 0
               for s in scalability]
    disagg_tp = [s["best_disagg"]["metrics_raw"]["throughput"] if s["best_disagg"] else 0
                 for s in scalability]

    # If all zero (SLO too strict), fall back to raw best throughput from all results
    if max(colo_tp) == 0 and max(disagg_tp) == 0:
        for i, g in enumerate(gpus):
            colo_best = max(
                (r for r in results if r.get("total_gpus") == g
                 and r.get("mode_label") == "colocated"),
                key=lambda r: r["metrics_raw"]["throughput"], default=None)
            disagg_best = max(
                (r for r in results if r.get("total_gpus") == g
                 and r.get("mode_label") == "disaggregated"),
                key=lambda r: r["metrics_raw"]["throughput"], default=None)
            colo_tp[i] = colo_best["metrics_raw"]["throughput"] if colo_best else 0
            disagg_tp[i] = disagg_best["metrics_raw"]["throughput"] if disagg_best else 0
        print("(SLO too strict — plotting raw best throughput per GPU count)")

    # Per-GPU efficiency
    colo_eff = [colo_tp[i] / gpus[i] if gpus[i] > 0 else 0
                for i in range(len(gpus))]
    disagg_eff = [disagg_tp[i] / gpus[i] if gpus[i] > 0 else 0
                  for i in range(len(gpus))]

    # Ideal linear scaling (from first GPU count with nonzero throughput)
    ideal = None
    for i, g in enumerate(gpus):
        if colo_tp[i] > 0 or disagg_tp[i] > 0:
            base_tp = max(colo_tp[i], disagg_tp[i])
            ideal = [base_tp * (gg / g) for gg in gpus]
            break

    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(14, 5.5))
    fig.suptitle("PD Disaggregation Scalability", fontsize=14, fontweight="bold", y=1.01)

    # ── Left: Throughput ──
    ax1.plot(gpus, colo_tp, "o-", color="#4472C4", linewidth=2.2, markersize=8,
             label="Colocated (best)")
    ax1.plot(gpus, disagg_tp, "s--", color="#C00000", linewidth=2.2, markersize=8,
             label="Disaggregated (best)")
    if ideal:
        ax1.plot(gpus, ideal, ":", color="gray", linewidth=1.5, alpha=0.7,
                 label=f"Ideal linear (×{gpus[0]} GPU baseline)")

    # Mark winner at each GPU count
    for i, g in enumerate(gpus):
        if colo_tp[i] == 0 and disagg_tp[i] == 0:
            continue
        winner = "C" if colo_tp[i] >= disagg_tp[i] else "D"
        y = max(colo_tp[i], disagg_tp[i])
        ax1.annotate(winner, (g, y), textcoords="offset points", xytext=(0, 10),
                     fontsize=10, fontweight="bold", ha="center",
                     color="#4472C4" if winner == "C" else "#C00000")

    ax1.set_xscale("log", base=2)
    ax1.set_xlabel("GPU Count (log₂)", fontsize=12)
    ax1.set_ylabel("Throughput (tok/s)", fontsize=12)
    ax1.legend(fontsize=10, loc="upper left")
    ax1.grid(True, alpha=0.3)
    ax1.set_xticks(gpus)
    ax1.set_xticklabels([str(g) for g in gpus])
    ax1.get_xaxis().set_major_formatter(mticker.ScalarFormatter())

    # ── Right: Per-GPU Efficiency ──
    ax2.plot(gpus, colo_eff, "o-", color="#4472C4", linewidth=2.2, markersize=8,
             label="Colocated")
    ax2.plot(gpus, disagg_eff, "s--", color="#C00000", linewidth=2.2, markersize=8,
             label="Disaggregated")

    ax2.set_xscale("log", base=2)
    ax2.set_xlabel("GPU Count (log₂)", fontsize=12)
    ax2.set_ylabel("Throughput per GPU (tok/s/GPU)", fontsize=12)
    ax2.legend(fontsize=10, loc="upper right")
    ax2.grid(True, alpha=0.3)
    ax2.set_xticks(gpus)
    ax2.set_xticklabels([str(g) for g in gpus])
    ax2.get_xaxis().set_major_formatter(mticker.ScalarFormatter())

    plt.tight_layout()

    png_path = os.path.splitext(xlsx_path)[0] + "_scalability.png"
    fig.savefig(png_path, dpi=150, bbox_inches="tight", facecolor="white")
    plt.close(fig)
    print(f"Scalability plot saved to: {png_path}")


def export_xlsx(results: list[dict], path: str) -> None:
    """Export results to XLSX with strategy_type, batch, and thr as separate columns."""
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "PD Sim Results"

    # Header styling
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    center_align = Alignment(horizontal="center")

    headers = [
        "strategy_type",
        "batch",
        "thr",
        "throughput_tok_s",
        "input_throughput_tok_s",
        "output_throughput_tok_s",
        "total_throughput_tok_s",
        "ttft_mean_ms", "ttft_p50_ms", "ttft_p90_ms", "ttft_p99_ms",
        "tpot_mean_ms", "tpot_p50_ms", "tpot_p90_ms", "tpot_p99_ms",
        "latency_p50_ms", "latency_p90_ms", "latency_p95_ms", "latency_p99_ms",
        "num_requests",
        "total_input_tokens",
        "total_output_tokens",
        "total_time_s",
        "cache_hit_rate",
        "attn_proj_pct", "ffn_proj_pct",
        "attn_prefill_pct", "attn_decode_pct",
        "rmsnorm_pct", "swiglu_pct", "rope_pct", "residual_add_pct",
        "lm_head_pct",
        "all_reduce_pct", "inter_stage_comm_pct",
        "kv_transfer_pct", "swap_pct",
        "score",
        "elapsed_s",
    ]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = center_align

    # Data rows
    for row, entry in enumerate(results, 2):
        m = entry["metrics_raw"]
        strategy_type, batch, thr = _parse_label(entry["label"])
        values = [
            strategy_type,
            batch,
            thr,
            m["throughput"],
            m["input_throughput"],
            m["output_throughput"],
            m["total_throughput"],
            m["mean_ttft_ms"], m["p50_ttft_ms"], m["p90_ttft_ms"], m["p99_ttft_ms"],
            m["mean_tpot_ms"], m["p50_tpot_ms"], m["p90_tpot_ms"], m["p99_tpot_ms"],
            m["p50_ms"], m["p90_ms"], m["p95_ms"], m["p99_ms"],
            m["num_requests"],
            m["total_input_tokens"],
            m["total_output_tokens"],
            m["total_time_s"],
            m.get("cache_hit_rate", 0.0),
            m.get("attn_proj_pct", 0.0), m.get("ffn_proj_pct", 0.0),
            m.get("attn_prefill_pct", 0.0), m.get("attn_decode_pct", 0.0),
            m.get("rmsnorm_pct", 0.0), m.get("swiglu_pct", 0.0),
            m.get("rope_pct", 0.0), m.get("residual_add_pct", 0.0),
            m.get("lm_head_pct", 0.0),
            m.get("all_reduce_pct", 0.0), m.get("inter_stage_comm_pct", 0.0),
            m.get("kv_transfer_pct", 0.0), m.get("swap_pct", 0.0),
            entry["score"],
            entry["elapsed"],
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            # Apply 3-decimal format to float columns (skip strategy_type, batch, thr)
            if isinstance(val, float):
                cell.number_format = '0.000'

    # Auto-width columns
    for col in range(1, len(headers) + 1):
        max_width = len(str(headers[col - 1]))
        for row in range(2, len(results) + 2):
            cell_val = str(ws.cell(row=row, column=col).value or "")
            max_width = max(max_width, min(len(cell_val), 40))
        ws.column_dimensions[get_column_letter(col)].width = max_width + 2

    # Freeze header row
    ws.freeze_panes = "A2"

    # ── Scalability summary sheet (if GPU sweep was run) ──
    _maybe_add_scalability_sheet(wb, results)

    wb.save(path)
    print(f"Results exported to: {path}")

    # ── Scalability plot ──
    _plot_scalability(results, path)
