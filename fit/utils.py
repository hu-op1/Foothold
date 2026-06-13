import json
import os
import numpy as np
from openpyxl import load_workbook


def _get(d, *keys):
    for k in keys:
        v = d.get(k)
        if v is not None:
            return v
    return None


def load_results(path):
    wb = load_workbook(path)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = dict(zip(headers, row))
        val = d.get("time_ms")
        if val is None or val == "OOM":
            continue
        d["time_ms"] = float(val)
        rows.append(d)
    return rows


def save_fitted_params(params, path):
    if os.path.isdir(path):
        path = os.path.join(path, "fitted_params.json")
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(params, f, indent=2)
    print(f"Fitted params saved to: {path}")


def load_fitted_params(path):
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


# ── roofline model ──────────────────────────────────────────────────────

def roofline_time(flops, bytes_moved, F_peak, B_peak, p):
    """Predicted time in seconds using smooth roofline model.

    p → 1: near-perfect overlap (worst-bound dominates weakly)
    p → ∞: zero overlap, approaches max(compute_time, memory_time)
    Typical p ∈ [1.5, 3] for real GPUs.
    """
    c = flops / F_peak
    m = bytes_moved / B_peak
    return (c ** p + m ** p) ** (1 / p)


def roofline_fit(flops_list, bytes_list, time_s_list):
    """Fit {F_peak, B_peak, p} from benchmark measurements.

    Args:
        flops_list: array-like, FLOP counts per data point
        bytes_list: array-like, bytes moved per data point
        time_s_list: array-like, measured times in seconds

    Returns:
        (F_peak, B_peak, p, r2)
    """
    from scipy.optimize import curve_fit

    flops = np.array(flops_list, dtype=np.float64)
    bytes_moved = np.array(bytes_list, dtype=np.float64)
    times = np.array(time_s_list, dtype=np.float64)

    def model(X, F, B, p):
        f, b_arr = X
        return (f / F + b_arr / B) ** (1 / p)

    # Initial guesses from raw FLOP/s and bytes/s ratios
    F0 = float(np.median(flops / times))
    B0 = float(np.median(bytes_moved / times))

    popt, _ = curve_fit(
        model,
        (flops, bytes_moved),
        times,
        p0=[F0 * 0.5, B0 * 0.5, 2.0],
        bounds=([1e9, 1e8, 1.0], [1e15, 1e13, 10.0]),
        maxfev=20000,
    )

    F_peak, B_peak, p = popt
    predicted = model((flops, bytes_moved), F_peak, B_peak, p)
    ss_res = float(np.sum((times - predicted) ** 2))
    ss_tot = float(np.sum((times - np.mean(times)) ** 2))
    r2 = 1 - ss_res / ss_tot if ss_tot > 0 else 0

    return F_peak, B_peak, p, r2
